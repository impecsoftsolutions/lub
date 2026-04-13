import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPABASE_URL = "https://qskziirjtzomrtckpzas.supabase.co"
ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InFza3ppaXJqdHpvbXJ0Y2twemFzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NTg1NjU4NjksImV4cCI6MjA3NDE0MTg2OX0.K888tIN7BzOraEQVXV6eTw5jQY2vyQSJZsGOrfUL89k"

HEADERS = {
    "apikey": ANON_KEY,
    "Authorization": f"Bearer {ANON_KEY}",
    "Content-Type": "application/json"
}

def fetch_all(endpoint, select, order="created_at.desc"):
    rows = []
    offset = 0
    batch = 1000
    while True:
        url = f"{SUPABASE_URL}/rest/v1/{endpoint}?select={select}&order={order}&limit={batch}&offset={offset}"
        r = requests.get(url, headers=HEADERS)
        r.raise_for_status()
        data = r.json()
        rows.extend(data)
        if len(data) < batch:
            break
        offset += batch
    return rows

users = fetch_all("users", "id,email,mobile_number,account_type,created_at,is_frozen")
print(f"Fetched {len(users)} users")

members = fetch_all("member_registrations", "email,full_name,company_name", "email.asc")
name_map = {m["email"]: (m.get("full_name", ""), m.get("company_name", "")) for m in members}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "LUB Users"

header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill   = PatternFill("solid", start_color="1F4E79")
header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin          = Side(style="thin", color="CCCCCC")
border        = Border(left=thin, right=thin, top=thin, bottom=thin)
alt_fill      = PatternFill("solid", start_color="EBF3FF")
data_font     = Font(name="Arial", size=10)
left_align    = Alignment(horizontal="left",   vertical="center")
center_align  = Alignment(horizontal="center", vertical="center")

columns = [
    ("S.No",         6),
    ("Full Name",   28),
    ("Email Address", 34),
    ("Mobile Number", 18),
    ("Company Name", 30),
    ("Account Type", 16),
    ("Status",       12),
    ("Registered On", 16),
]

for ci, (hdr, w) in enumerate(columns, 1):
    cell = ws.cell(row=1, column=ci, value=hdr)
    cell.font       = header_font
    cell.fill       = header_fill
    cell.alignment  = header_align
    cell.border     = border
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.row_dimensions[1].height = 30

account_labels = {
    "member":       "Member",
    "admin":        "Admin",
    "both":         "Member + Admin",
    "general_user": "General User",
}

for ri, user in enumerate(users, 2):
    fill       = alt_fill if ri % 2 == 0 else None
    email      = user.get("email", "")
    name, company = name_map.get(email, ("", ""))
    joined     = user.get("created_at", "")[:10]
    acc_type   = account_labels.get(user.get("account_type", ""), user.get("account_type", ""))
    status     = "Frozen" if user.get("is_frozen") else "Active"

    row_data = [ri - 1, name, email, user.get("mobile_number", "") or "", company, acc_type, status, joined]

    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.font   = data_font
        cell.border = border
        if fill:
            cell.fill = fill
        cell.alignment = center_align if ci in (1, 6, 7, 8) else left_align

    ws.row_dimensions[ri].height = 18

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

sr = len(users) + 3
ws.cell(row=sr, column=1, value="Total Users:").font = Font(name="Arial", bold=True, size=10)
ws.cell(row=sr, column=2, value=f"=COUNTA(C2:C{len(users)+1})").font = Font(name="Arial", bold=True, size=10)

out = r"C:\webprojects\lub\LUB_Users_Export.xlsx"
wb.save(out)
print(f"Saved: {out}")
