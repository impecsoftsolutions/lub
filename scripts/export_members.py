import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPABASE_URL = "https://qskziirjtzomrtckpzas.supabase.co"
ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InFza3ppaXJqdHpvbXJ0Y2twemFzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NTg1NjU4NjksImV4cCI6MjA3NDE0MTg2OX0.K888tIN7BzOraEQVXV6eTw5jQY2vyQSJZsGOrfUL89k"

headers = {
    "apikey": ANON_KEY,
    "Authorization": f"Bearer {ANON_KEY}",
    "Content-Type": "application/json"
}

all_rows = []
batch_size = 1000
offset = 0

while True:
    url = (
        f"{SUPABASE_URL}/rest/v1/member_registrations"
        f"?select=full_name,email,mobile_number,company_name,status,created_at"
        f"&order=full_name.asc"
        f"&limit={batch_size}&offset={offset}"
    )
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    batch = resp.json()
    if not batch:
        break
    all_rows.extend(batch)
    if len(batch) < batch_size:
        break
    offset += batch_size

print(f"Fetched {len(all_rows)} records")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "LUB Members"

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="1F4E79")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

alt_fill = PatternFill("solid", start_color="EBF3FF")

columns = [
    ("S.No", 6),
    ("Full Name", 28),
    ("Email Address", 34),
    ("Mobile Number", 18),
    ("Company Name", 32),
    ("Status", 14),
    ("Joined On", 16),
]

for col_idx, (header, width) in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 30

data_font = Font(name="Arial", size=10)
data_align_left = Alignment(horizontal="left", vertical="center")
data_align_center = Alignment(horizontal="center", vertical="center")

for row_idx, rec in enumerate(all_rows, start=2):
    fill = alt_fill if row_idx % 2 == 0 else None
    joined = rec.get("created_at", "")[:10] if rec.get("created_at") else ""

    row_data = [
        row_idx - 1,
        rec.get("full_name", ""),
        rec.get("email", ""),
        rec.get("mobile_number", ""),
        rec.get("company_name", ""),
        rec.get("status", "").capitalize(),
        joined,
    ]

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = border
        if fill:
            cell.fill = fill
        cell.alignment = data_align_center if col_idx in (1, 6, 7) else data_align_left

    ws.row_dimensions[row_idx].height = 18

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

summary_row = len(all_rows) + 3
ws.cell(row=summary_row, column=1, value="Total Members:").font = Font(name="Arial", bold=True, size=10)
ws.cell(row=summary_row, column=2, value=f'=COUNTA(B2:B{len(all_rows)+1})').font = Font(name="Arial", bold=True, size=10)

out_path = r"C:\webprojects\lub\LUB_Members_Export.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
